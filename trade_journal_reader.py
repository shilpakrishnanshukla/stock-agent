"""
trade_journal_reader.py
Reads currently OPEN positions from the "Trade Journal" tab and returns them
in the shape analyze.py's holdings pipeline expects:
    [{"ticker": "AAPL", "shares": 10, "cost_basis": 187.32}, ...]

Column layout matches the actual "Trade Journal" tab header row (row 4):
Entry Date(A), Company(B), Ticker(C), Entry Price(D), Shares(E), Position
Value(F), Stop Loss(G), Target(H), ... Exit Date(M), ... Status(U), ...

There is no dedicated "Market" column in this workbook - SG (SGX) tickers
are identified by a ".SI" suffix (e.g. "D05.SI"); everything else is
treated as US.

A row counts as open when its Status column says "Open" (case-
insensitive). If Status is blank/unrecognized, an empty Exit Date is used
as the fallback signal. Rows explicitly marked "Closed" are always
excluded even if Exit Date happens to be blank.

Multiple open lots of the same ticker are combined into one holding, with
cost_basis computed as the shares-weighted average entry price - so two
partial buys of the same name show up as a single, correctly-averaged
position rather than two separate rows.

Excel is treated as the sole source of truth for holdings: this replaces
portfolio.json's "holdings" list entirely rather than merging with it.
"""

from collections import defaultdict

import openpyxl

SHEET_NAME = "Trade Journal"
FIRST_DATA_ROW = 5

# 1-indexed column numbers, per the actual Trade Journal header row (row 4)
COL_TICKER = 3        # C
COL_ENTRY_PRICE = 4   # D
COL_SHARES = 5        # E
COL_EXIT_DATE = 13    # M
COL_STATUS = 21       # U


def _is_open(status_value, exit_date_value) -> bool:
    if isinstance(status_value, str):
        s = status_value.strip().lower()
        if s == "open":
            return True
        if s == "closed":
            return False
    # Status blank/unrecognized - fall back to whether an exit date was entered.
    return exit_date_value in (None, "")


def _infer_market(ticker: str) -> str:
    """No dedicated Market column exists - SG names carry a .SI suffix,
    everything else is treated as US."""
    return "SG" if ticker.upper().endswith(".SI") else "US"


def read_holdings_from_trade_journal(local_path: str, market: str = None, sheet_name: str = SHEET_NAME) -> list:
    """
    market: if given (e.g. "US" or "SG"), only tickers matching that
    inferred market are included - so analyze.py and analyze_sg.py each
    pick up only their own holdings from the same shared workbook.
    """
    wb = openpyxl.load_workbook(local_path, data_only=True)
    ws = wb[sheet_name]

    lots_by_ticker = defaultdict(lambda: {"shares": 0.0, "cost_total": 0.0})

    row = FIRST_DATA_ROW
    while ws.cell(row=row, column=COL_TICKER).value not in (None, ""):
        ticker = ws.cell(row=row, column=COL_TICKER).value
        entry_price = ws.cell(row=row, column=COL_ENTRY_PRICE).value
        shares = ws.cell(row=row, column=COL_SHARES).value
        exit_date = ws.cell(row=row, column=COL_EXIT_DATE).value
        status = ws.cell(row=row, column=COL_STATUS).value
        row += 1

        if not ticker or not entry_price or not shares:
            continue

        key = str(ticker).strip().upper()

        if market and _infer_market(key) != market.strip().upper():
            continue
        if not _is_open(status, exit_date):
            continue

        lot = lots_by_ticker[key]
        lot["shares"] += float(shares)
        lot["cost_total"] += float(shares) * float(entry_price)

    holdings = []
    for ticker, lot in lots_by_ticker.items():
        if lot["shares"] <= 0:
            continue
        shares = lot["shares"]
        holdings.append({
            "ticker": ticker,
            "shares": int(shares) if shares == int(shares) else round(shares, 4),
            "cost_basis": round(lot["cost_total"] / shares, 4),
        })
    return holdings
