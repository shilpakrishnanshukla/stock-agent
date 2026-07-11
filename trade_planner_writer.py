"""
trade_planner_writer.py
Appends new candidate rows to the "Trade Planner" tab of the trading
workbook, without touching any row that already exists (so your Status /
Decision Notes annotations are never overwritten).

Formula columns (I, P, Q, R, S, T, U, V, W, X, Z, AA) are copied down from
row 5 using openpyxl's Translator, so Settings!$B$4-style absolute refs stay
pinned while row-relative refs (J5, M5, etc.) shift to the new row.

Usage:
    from trade_planner_writer import update_trade_planner
    update_trade_planner("workbook.xlsx", candidates)

Each item in `candidates` is a dict with these keys:
    ticker, company, market, setup_type,
    trend_score, momentum_score, earnings_score, location_score,
    entry, support, atr, stop, resistance, target,
    fee_estimate           (optional, defaults to 2.2)
    status                 (optional, defaults to "New")
    notes                  (optional, defaults to "Auto-added <date>")
"""

from datetime import date

import openpyxl
from openpyxl.formula.translate import Translator

SHEET_NAME = "Trade Planner"
HEADER_ROW = 4
FIRST_DATA_ROW = 5

# Manual-input columns (letters) mapped to candidate dict keys
FIELD_COLUMNS = {
    "A": "ticker",
    "B": "company",
    "C": "market",
    "D": "setup_type",
    "E": "trend_score",
    "F": "momentum_score",
    "G": "earnings_score",
    "H": "location_score",
    "J": "entry",
    "K": "support",
    "L": "atr",
    "M": "stop",
    "N": "resistance",
    "O": "target",
    "Y": "fee_estimate",
    "AB": "status",
    "AC": "notes",
}

# Formula columns to copy down from the template row
FORMULA_COLUMNS = ["I", "P", "Q", "R", "S", "T", "U", "V", "W", "X", "Z", "AA"]


def _find_last_row(ws) -> int:
    row = FIRST_DATA_ROW
    last = FIRST_DATA_ROW - 1
    while ws.cell(row=row, column=1).value not in (None, ""):
        last = row
        row += 1
    return last


def _existing_tickers(ws, last_row: int) -> set:
    tickers = set()
    for r in range(FIRST_DATA_ROW, last_row + 1):
        v = ws.cell(row=r, column=1).value
        if v:
            tickers.add(str(v).strip().upper())
    return tickers


def update_trade_planner(local_path: str, candidates: list, sheet_name: str = SHEET_NAME) -> int:
    """
    Appends any candidates not already present (by ticker) below the last
    used row in Trade Planner. Returns the number of rows added.
    """
    wb = openpyxl.load_workbook(local_path, data_only=False)
    ws = wb[sheet_name]

    last_row = _find_last_row(ws)
    already_have = _existing_tickers(ws, last_row)

    template_row = FIRST_DATA_ROW  # row 5 holds the formula pattern
    next_row = last_row + 1
    added = 0

    for c in candidates:
        ticker = str(c.get("ticker", "")).strip().upper()
        if not ticker or ticker in already_have:
            continue

        # Manual-input fields
        for col_letter, key in FIELD_COLUMNS.items():
            if key == "status":
                value = c.get("status", "New")
            elif key == "notes":
                value = c.get("notes", f"Auto-added by daily pipeline - {date.today().isoformat()}")
            elif key == "fee_estimate":
                value = c.get("fee_estimate", 2.2)
            else:
                value = c.get(key)
            ws[f"{col_letter}{next_row}"] = value

        # Formula fields, translated from the template row
        for col_letter in FORMULA_COLUMNS:
            template_formula = ws[f"{col_letter}{template_row}"].value
            if isinstance(template_formula, str) and template_formula.startswith("="):
                new_formula = Translator(
                    template_formula, origin=f"{col_letter}{template_row}"
                ).translate_formula(f"{col_letter}{next_row}")
                ws[f"{col_letter}{next_row}"] = new_formula

        already_have.add(ticker)
        next_row += 1
        added += 1

    wb.save(local_path)
    return added
